import csv
import openpyxl as op


def main():
    Datas = []
    file = input("Enter the excel File Location Path: ")
    wb = op.load_workbook(file)
    shnames = wb.sheetnames
    print(wb.sheetnames)    
    sheet_select = input("Enter the sheet 0,1,2.. :")
    colu = input("Enter the Columns eg. 1,2 where A = 1 : ").split(",")
    ws = wb[shnames[int(sheet_select)]]
    FileName = input("Enter the file name for Output file .csv: ")
    v = [[] for x in range(len(colu))]


    for i in range(len(colu)):
        for row in ws.iter_cols(min_row=1, min_col=int(colu[i]), max_col=int(colu[i]), values_only=True): 
            for cell in row:
                v[i].append(cell)

    list_all   = list(zip(v[0], v[1]))
    try:
        open(f"{FileName}", "x")
    except FileExistsError:
        pass

    with open(f"{FileName}","w") as file:
        csvwriter = csv.writer(file,lineterminator="\n")
        csvwriter.writerows(list_all)


if __name__=="__main__":
    main()
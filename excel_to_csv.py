import csv
import sys
from typing import List, Tuple, Any, Optional

# Third-party libraries
try:
    import openpyxl as op
    from openpyxl.utils import column_index_from_string, get_column_letter
except ImportError:
    print("Error: Required libraries (openpyxl) are not installed.")
    print("Please run: pip install openpyxl")
    sys.exit(1)


# --- Data Structures and Utilities ---

def get_valid_column_indices(col_input: str) -> Optional[List[int]]:
    """
    Validates and converts user input (e.g., '1,2' or 'A,B') into a list of 1-based column indices.
    """
    column_indices = []
    
    # Standardize input: convert all to uppercase and remove spaces
    col_input = col_input.strip().upper().replace(" ", "")

    if not col_input:
        print("Error: Column input cannot be empty.")
        return None

    try:
        raw_cols = col_input.split(',')
        
        for col in raw_cols:
            if not col: continue # Skip empty strings from multiple commas

            # Check if input is a letter (e.g., 'A', 'B')
            if col.isalpha():
                index = column_index_from_string(col)
                column_indices.append(index)
            # Check if input is a number (e.g., '1', '2')
            elif col.isdigit():
                index = int(col)
                if index < 1:
                    print(f"Error: Column number must be 1 or greater, not {index}.")
                    return None
                column_indices.append(index)
            else:
                print(f"Error: Invalid column format '{col}'. Use '1,2' or 'A,B'.")
                return None
                
        if not column_indices:
             print("Error: No valid columns were specified.")
             return None
             
        return column_indices
        
    except Exception as e:
        print(f"An unexpected error occurred during column parsing: {e}")
        return None

def extract_excel_columns(ws: op.worksheet.worksheet.Worksheet, col_indices: List[int]) -> Optional[List[Tuple[Any, ...]]]:
    """
    Extracts the specified columns from the worksheet and zips them into rows.
    """
    try:
        # Create a list of lists, where each inner list will hold data for one column
        columns_data = [[] for _ in col_indices]
        
        # Iterate over the target columns
        for i, col_index in enumerate(col_indices):
            
            # Use iter_cols for clean column-wise reading
            # min_row=1 ensures we start at the top
            for row in ws.iter_cols(min_row=1, min_col=col_index, max_col=col_index, values_only=True):
                # 'row' here is a tuple containing all cell values for that column
                for cell_value in row:
                    columns_data[i].append(cell_value)
        
        # Zip the individual columns into a list of rows (tuples)
        # Example: [[A1, A2], [B1, B2]] -> [(A1, B1), (A2, B2)]
        if not columns_data or not columns_data[0]:
            print("Warning: The selected columns appear to be empty or contain no data.")
            return []
            
        list_of_rows = list(zip(*columns_data))
        return list_of_rows

    except Exception as e:
        print(f"An error occurred during data extraction: {e}")
        return None

def write_to_csv(file_name: str, data: List[Tuple[Any, ...]]) -> bool:
    """Writes the list of row tuples to the specified CSV file."""
    try:
        # 'w' mode will create the file if it doesn't exist, and overwrite if it does
        with open(file_name, "w", newline="", encoding='utf-8') as file:
            csvwriter = csv.writer(file, lineterminator="\n")
            csvwriter.writerows(data)
        return True
    except IOError as e:
        print(f"Error writing to CSV file '{file_name}': {e}")
        return False
    except Exception as e:
        print(f"An unexpected error occurred during CSV writing: {e}")
        return False

# --- Main Execution ---

def main():
    """The main function controlling the program flow."""
    print("✨ Professional Excel Column Extractor to CSV Converter ✨")
    
    # 1. Get File Input and Load Workbook
    file_path = input("\nEnter the Excel File Location Path: ").strip()
    
    try:
        wb = op.load_workbook(file_path)
    except FileNotFoundError:
        print(f"Error: File not found at '{file_path}'. Exiting.")
        return
    except op.utils.exceptions.InvalidFileException:
        print(f"Error: '{file_path}' is not a valid Excel file.")
        return
    except Exception as e:
        print(f"An unexpected error occurred while loading the workbook: {e}")
        return

    shnames = wb.sheetnames
    print(f"\nAvailable sheets: {shnames}")
    
    # 2. Select Sheet
    try:
        sheet_input = input("Enter the sheet name or its 0-based index (e.g., 'Sheet1' or '0'): ").strip()
        
        if sheet_input.isdigit():
            sheet_index = int(sheet_input)
            ws = wb[shnames[sheet_index]]
        else:
            ws = wb[sheet_input] # Direct sheet name lookup
        
        print(f"✅ Selected sheet: '{ws.title}' (Max columns: {ws.max_column})")
        
    except (IndexError, KeyError):
        print(f"Error: Invalid sheet name or index '{sheet_input}'. Exiting.")
        return
    except Exception as e:
        print(f"An unexpected error occurred during sheet selection: {e}")
        return

    # 3. Get and Validate Columns
    col_input = input("Enter the columns to extract (e.g., '1,2,5' or 'A,B,E'): ").strip()
    col_indices = get_valid_column_indices(col_input)
    
    if not col_indices:
        return
    
    # Convert indices back to letters for user feedback
    col_letters = ", ".join(get_column_letter(i) for i in col_indices)
    print(f"✅ Columns selected: {col_letters} (Indices: {col_indices})")

    # 4. Extract Data
    list_of_rows = extract_excel_columns(ws, col_indices)
    
    if list_of_rows is None: # Critical extraction error
        print("Data extraction failed. Exiting.")
        return
        
    if not list_of_rows: # Empty data
        print("No data was extracted from the selected columns. Exiting.")
        return

    print(f"Extracted {len(list_of_rows)} rows of data.")

    # 5. Write to CSV
    output_filename = input("Enter the desired file name for the Output CSV (e.g., output.csv): ").strip()
    
    if not output_filename.lower().endswith('.csv'):
        output_filename += '.csv'
        
    if write_to_csv(output_filename, list_of_rows):
        print(f"\n🎉 Success! Data saved to '{output_filename}'.")
    else:
        print("\nOperation failed to complete successfully.")


if __name__ == "__main__":
    main()